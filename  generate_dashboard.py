#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime
import os

def read_excel(file_path):
    """Читає Excel і повертає дані"""
    wb = openpyxl.load_workbook(file_path)
    
    люди = []
    вакансії = []
    
    # Читаємо Люди
    if 'Люди' in wb.sheetnames:
        ws = wb['Люди']
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row[1]:  # Якщо ім'я не пусто
                try:
                    люди.append({
                        'ім\'я': str(row[1]) if row[1] else '',
                        'посада': str(row[2]) if row[2] else '',
                        'відділ': str(row[5]) if len(row) > 5 and row[5] else '',
                        'статус': str(row[6]) if len(row) > 6 and row[6] else '',
                        'дата_початку': row[3],
                        'дата_завершення_випробування': row[4] if len(row) > 4 else '',
                        'зп': float(row[8]) if len(row) > 8 and isinstance(row[8], (int, float)) else 0,
                        'кпі': float(row[9]) if len(row) > 9 and isinstance(row[9], (int, float)) else 0,
                        'performance': row[10] if len(row) > 10 else 0,
                        'impact': row[11] if len(row) > 11 else 0,
                        'ризик': str(row[12]) if len(row) > 12 and row[12] else ''
                    })
                except Exception as e:
                    print(f"Помилка рядку {idx}: {e}")
                    continue
    
    # Читаємо Вакансії
    if 'Вакансії' in wb.sheetnames:
        ws = wb['Вакансії']
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row[1]:
                try:
                    вакансії.append({
                        'назва': str(row[1]) if row[1] else '',
                        'керівник': str(row[2]) if row[2] else '',
                        'дата_відкриття': row[3],
                        'дата_закриття': row[4] if len(row) > 4 else '',
                        'статус': str(row[5]) if len(row) > 5 and row[5] else '',
                        'ціна': float(row[12]) if len(row) > 12 and isinstance(row[12], (int, float)) else 0
                    })
                except Exception as e:
                    print(f"Помилка рядку вакансії {idx}: {e}")
                    continue
    
    return люди, вакансії

def generate_html(люди, вакансії):
    """Генерує HTML дашборд"""
    
    now = datetime.now()
    active = [p for p in люди if p['статус'] == 'Працює']
    probation = [p for p in люди if p['статус'] == 'Випробуванн']
    total = len(active) + len(probation)
    
    # Розрахунки
    fop = sum(p['зп'] + p['кпі'] for p in active)
    open_vacs = len([v for v in вакансії if v['статус'] == 'Відкрита'])
    
    # Алерти
    alerts = []
    if probation:
        alerts.append(f"🔴 {len(probation)} людей на випробуванні")
    if open_vacs > 0:
        alerts.append(f"🟡 {open_vacs} вакансій відкритих")
    if not alerts:
        alerts.append("✅ Все стабільно")
    
    alerts_html = ''.join([f'<div class="alert">{a}</div>' for a in alerts])
    
    # Генеруємо HTML
    html = f'''<!DOCTYPE html>
<html lang="uk">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HR Dashboard</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#f5f4f1;color:#1a1a18;font-family:system-ui,sans-serif;font-size:14px;line-height:1.5}}

.login-container{{position:fixed;inset:0;background:linear-gradient(135deg,#378ADD,#1D9E75);display:flex;align-items:center;justify-content:center;z-index:1000}}
.login-box{{background:#fff;border-radius:10px;padding:2rem;box-shadow:0 10px 40px rgba(0,0,0,.2);width:100%;max-width:320px}}
.login-box h1{{font-size:20px;margin-bottom:1.5rem;text-align:center;color:#1a1a18}}
.login-box input{{width:100%;padding:10px;border:.5px solid #d1d5db;border-radius:7px;margin-bottom:1rem;font-size:14px;font-family:inherit}}
.login-box button{{width:100%;padding:10px;background:#378ADD;color:#fff;border:none;border-radius:7px;font-weight:500;cursor:pointer;font-size:14px}}
.login-box button:hover{{background:#1D9E75}}
.login-error{{color:#E24B4A;font-size:12px;margin-bottom:1rem;text-align:center;display:none}}
.login-container.hidden{{display:none}}

.layout{{max-width:1280px;margin:0 auto;padding:2rem 1.5rem}}
header{{margin-bottom:2rem;display:flex;justify-content:space-between;align-items:flex-start}}
header h1{{font-size:24px;font-weight:500}}
.logout{{font-size:12px;background:#f0efe9;border:.5px solid #d1d5db;border-radius:100px;padding:6px 12px;color:#5f5e5a;cursor:pointer;font-family:inherit;border-style:solid}}
.logout:hover{{background:#d1d5db}}
.updated{{font-size:11px;background:#f0efe9;border:.5px solid #d1d5db;border-radius:100px;padding:4px 12px;color:#5f5e5a}}

.sec{{font-size:11px;font-weight:500;letter-spacing:.08em;color:#8a8980;text-transform:uppercase;margin:2rem 0 .75rem}}
.metric-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:2rem}}
.metric{{background:#f0efe9;border-radius:7px;padding:1rem;text-align:center}}
.metric .lbl{{font-size:12px;color:#5f5e5a;margin-bottom:4px}}
.metric .val{{font-size:28px;font-weight:500}}
.metric .sub{{font-size:11px;color:#8a8980;margin-top:4px}}
.metric.alert .val{{color:#E24B4A}}

.card{{background:#fff;border:.5px solid rgba(0,0,0,.09);border-radius:10px;padding:1.25rem;margin-bottom:1rem}}
.two-grid{{display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-bottom:1rem}}
.row{{display:flex;justify-content:space-between;padding:8px 0;border-bottom:.5px solid rgba(0,0,0,.09);font-size:13px}}
.row:last-child{{border-bottom:none}}
.row-label{{color:#5f5e5a}}
.row-value{{color:#1a1a18;font-weight:500}}
.alert{{padding:10px;background:#FEE2E2;border-left:3px solid #E24B4A;border-radius:7px;margin-bottom:8px;font-size:12px;color:#991B1B}}
footer{{text-align:center;font-size:11px;color:#8a8980;margin-top:3rem;padding-top:1.5rem;border-top:.5px solid rgba(0,0,0,.09)}}
</style>
</head>
<body>

<div class="login-container" id="login">
  <div class="login-box">
    <h1>HR Dashboard</h1>
    <div class="login-error" id="error">Неправильний пароль</div>
    <input type="password" id="password" placeholder="Пароль" onkeypress="if(event.key==='Enter') login()">
    <button onclick="login()">Вхід</button>
  </div>
</div>

<div class="layout" id="app" style="display:none">
  <header>
    <div><h1>HR Dashboard</h1><p style="color:#5f5e5a;margin-top:3px">Власники</p></div>
    <div style="display:flex;gap:1rem;align-items:center">
      <span class="updated" id="updated">{now.strftime('%d.%m.%Y')}</span>
      <button class="logout" onclick="logout()">Вихід</button>
    </div>
  </header>

  <div class="sec">Ключові показники</div>
  <div class="metric-grid">
    <div class="metric"><div class="lbl">Людей</div><div class="val">{total}</div></div>
    <div class="metric"><div class="lbl">На випробуванні</div><div class="val">{len(probation)}</div></div>
    <div class="metric alert"><div class="lbl">ФОП</div><div class="val">{int(fop):,}</div><div class="sub">грн/м</div></div>
    <div class="metric"><div class="lbl">Вакансій</div><div class="val">{open_vacs}</div></div>
  </div>

  <div class="sec">Потребує уваги</div>
  <div class="card">{alerts_html}</div>

  <div class="two-grid">
    <div>
      <div class="sec" style="margin-top:0">Команда</div>
      <div class="card" style="margin-bottom:0">
        <div class="row"><span class="row-label">Працює</span><span class="row-value">{len(active)}</span></div>
        <div class="row"><span class="row-label">Випробування</span><span class="row-value">{len(probation)}</span></div>
      </div>
    </div>
    <div>
      <div class="sec" style="margin-top:0">Рекрутинг</div>
      <div class="card" style="margin-bottom:0">
        <div class="row"><span class="row-label">Відкриті</span><span class="row-value">{open_vacs}</span></div>
        <div class="row"><span class="row-label">Закрито</span><span class="row-value">{len([v for v in вакансії if v['статус'] == 'Закрита'])}</span></div>
      </div>
    </div>
  </div>

  <footer>Оновлено автоматично</footer>
</div>

<script>
const PASSWORD = 'настя 25';

function login() {{
  if (document.getElementById('password').value === PASSWORD) {{
    localStorage.setItem('auth', 'true');
    document.getElementById('login').classList.add('hidden');
    document.getElementById('app').style.display = '';
  }} else {{
    document.getElementById('error').style.display = 'block';
    document.getElementById('password').value = '';
  }}
}}

function logout() {{
  localStorage.removeItem('auth');
  location.reload();
}}

if (localStorage.getItem('auth') === 'true') {{
  document.getElementById('login').classList.add('hidden');
  document.getElementById('app').style.display = '';
}}
</script>

</body>
</html>'''
    
    return html

if __name__ == '__main__':
    excel_path = 'HR_System.xlsx'
    if not os.path.exists(excel_path):
        excel_path = os.path.join(os.path.dirname(__file__), 'HR_System.xlsx')
    
    print(f"Читаємо: {excel_path}")
    люди, вакансії = read_excel(excel_path)
    
    print(f"✅ Прочитано людей: {len(люди)}")
    print(f"✅ Прочитано вакансій: {len(вакансії)}")
    
    html = generate_html(люди, вакансії)
    
    output_path = 'index.html'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ Дашборд сгенеровано: {output_path}")
